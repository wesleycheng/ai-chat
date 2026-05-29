import os
import uuid
from datetime import datetime
from typing import List, Optional

import aiofiles
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..core.database import get_db
from ..core.security import get_current_user
from ..models.models import File as DBFile, FileStatus, User
from ..schemas.schemas import FileCreate, FileResponse, FileListResponse

router = APIRouter(prefix="/files", tags=["files"])

# 支持的文件扩展名和对应的MIME类型
ALLOWED_EXTENSIONS = {
    ".pdf": ["application/pdf"],
    ".docx": ["application/vnd.openxmlformats-officedocument.wordprocessingml.document"],
    ".xlsx": ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"],
    ".txt": ["text/plain"],
    ".md": ["text/markdown", "text/plain"],
    ".jpg": ["image/jpeg"],
    ".jpeg": ["image/jpeg"],
    ".png": ["image/png"],
}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


def get_extension_from_filename(filename: str) -> Optional[str]:
    """从文件名获取扩展名"""
    if not filename:
        return None
    ext = os.path.splitext(filename)[1].lower()
    return ext if ext in ALLOWED_EXTENSIONS else None


async def validate_file(file: UploadFile) -> str:
    """验证文件类型和大小，返回扩展名"""
    # 检查文件大小
    contents = await file.read(MAX_FILE_SIZE + 1)
    await file.seek(0)
    
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"文件大小超过 {MAX_FILE_SIZE / 1024 / 1024}MB 限制")
    
    # 先尝试从文件名获取扩展名
    ext = get_extension_from_filename(file.filename)
    
    # 如果是文本文件，直接允许（不需要检查魔数）
    if ext in [".txt", ".md"]:
        return ext
    
    # 其他文件类型检查MIME类型
    if file.content_type:
        allowed_mimes = ALLOWED_EXTENSIONS.get(ext, [])
        if file.content_type not in allowed_mimes and ext:
            # 有些浏览器可能发送不同的MIME类型，用扩展名判断
            pass
    
    if not ext:
        raise HTTPException(status_code=400, detail="不支持的文件类型，支持：pdf/docx/xlsx/txt/md/jpg/png")
    
    return ext


async def parse_file_content(file_path: str, extension: str) -> Optional[str]:
    """解析文件内容为纯文本"""
    ext = extension.lower()
    
    try:
        if ext == ".pdf":
            return await parse_pdf(file_path)
        elif ext in [".docx", ".doc"]:
            return await parse_docx(file_path)
        elif ext in [".xlsx", ".xls"]:
            return await parse_xlsx(file_path)
        elif ext in [".txt", ".md"]:
            async with aiofiles.open(file_path, "r", encoding="utf-8") as f:
                return await f.read()
        elif ext in [".jpg", ".jpeg", ".png"]:
            return "[图片文件，暂不支持OCR解析]"
    except Exception as e:
        return f"[文件解析失败: {str(e)}]"
    
    return None


async def parse_pdf(file_path: str) -> str:
    """解析 PDF"""
    import pymupdf
    text_parts = []
    with pymupdf.open(file_path) as doc:
        for page in doc:
            text_parts.append(page.get_text())
    return "\n".join(text_parts)


async def parse_docx(file_path: str) -> str:
    """解析 DOCX"""
    from docx import Document
    doc = Document(file_path)
    return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])


async def parse_xlsx(file_path: str) -> str:
    """解析 XLSX"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True)
    parts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                parts.append(" | ".join([str(cell) if cell else "" for cell in row]))
    return "\n".join(parts)


@router.post("/upload", response_model=FileResponse)
async def upload_file(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """上传文件并立即解析内容"""
    # 验证文件
    try:
        ext = await validate_file(file)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    # 生成文件路径
    file_id = str(uuid.uuid4())
    upload_dir = os.getenv("UPLOAD_DIR", "./uploads")
    file_path = os.path.join(upload_dir, f"{file_id}{ext}")
    
    # 确保目录存在
    os.makedirs(upload_dir, exist_ok=True)
    
    # 保存文件
    contents = await file.read()
    async with aiofiles.open(file_path, "wb") as f:
        await f.write(contents)
    
    # 立即解析文件内容
    content_text = await parse_file_content(file_path, ext)
    
    # 创建数据库记录
    db_file = DBFile(
        id=file_id,
        user_id=current_user.id,
        filename=file.filename or "unknown",
        file_ext=ext,
        file_size=len(contents),
        file_path=file_path,
        parse_status=FileStatus.COMPLETED if content_text and not content_text.startswith("[") else FileStatus.FAILED,
        content_text=content_text,
    )
    db.add(db_file)
    await db.commit()
    await db.refresh(db_file)
    
    return db_file


@router.get("", response_model=FileListResponse)
async def list_files(
    skip: int = 0,
    limit: int = 20,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取文件列表"""
    stmt = select(DBFile).where(DBFile.user_id == current_user.id).order_by(DBFile.created_at.desc())
    stmt = stmt.offset(skip).limit(limit)
    result = await db.execute(stmt)
    files = result.scalars().all()
    
    count_stmt = select(DBFile).where(DBFile.user_id == current_user.id)
    count_result = await db.execute(count_stmt)
    total = len(count_result.scalars().all())
    
    return {"items": files, "total": total}


@router.get("/{file_id}", response_model=FileResponse)
async def get_file(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取文件详情"""
    stmt = select(DBFile).where(DBFile.id == file_id, DBFile.user_id == current_user.id)
    result = await db.execute(stmt)
    file = result.scalar_one_or_none()
    
    if not file:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    return file


@router.delete("/{file_id}")
async def delete_file(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """删除文件"""
    stmt = select(DBFile).where(DBFile.id == file_id, DBFile.user_id == current_user.id)
    result = await db.execute(stmt)
    file = result.scalar_one_or_none()
    
    if not file:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    # 删除物理文件
    if file.file_path and os.path.exists(file.file_path):
        os.remove(file.file_path)
    
    # 删除数据库记录
    await db.delete(file)
    await db.commit()
    
    return {"message": "删除成功"}
